
try:
    import cupy as np
    print(f"cupy{np.__version__}")
except ModuleNotFoundError:
    print(ModuleNotFoundError)
    import numpy as np
import pandas as pd
from openpyxl import load_workbook

def Fun_diffraction(phi, matrixband, matrixPhaseBandGap, wavelennum, samplenum, Dx, T, N, bandnum, wavelen0, lam):

    phase = np.zeros([2 * N + 1, 2 * N + 1])
    phase[bandnum >= 1] = phi[bandnum[bandnum >= 1] - 1] + matrixPhaseBandGap[bandnum >= 1]
    num1 = np.arange(samplenum-1, -1, -1)
    yy = Dx * np.ones([samplenum, 1]) * (num1 - samplenum / 2 + 0.5)
    YY = yy.T
    num2 = np.arange(samplenum)
    XX = Dx * np.ones([samplenum, 1]) * (num2 - samplenum / 2 + 0.5)

    BY = np.zeros([samplenum, samplenum]).astype(int)
    BY[YY >= 0] = N - np.ceil((YY[YY >= 0] - T / 2) / T)
    BY[YY < 0] = N + np.ceil((np.abs(YY[YY < 0]) - T / 2) / T)
    BX = np.zeros([samplenum, samplenum]).astype(int)
    BX[XX >= 0] = N + np.ceil((XX[XX >= 0] - T / 2) / T)
    BX[XX < 0] = N - np.ceil((np.abs(XX[XX < 0]) - T / 2) / T)
    PhaseSamplePoints = phase[BX, BY]
    AmpSamplePoints = matrixband[BX, BY]
    Ex0 = AmpSamplePoints * np.exp(1j * PhaseSamplePoints)
    Ey0 = AmpSamplePoints * np.exp(1j * (PhaseSamplePoints + np.pi / 2))
    ratio1, distance1, strength1, SL, efficiency = AngularSpectrumDiffraction(Ex0, Ey0, wavelennum, samplenum, Dx, wavelen0, lam)
    return ratio1, distance1, strength1, SL, efficiency

def AngularSpectrumDiffraction(Ex0, Ey0, wavelennum, samplenum, Dx, wavelen0, lam):

    global Zd1
    n_refra = 1
    nn = 250
    # Num = 526
    # Zd1 = np.arange(269, 374 + 0.2, 0.2)*wavelen0
    Num = 571
    # Zd1 = np.arange(430, 450 + 0.5, 0.5) * wavelen0
    Zd1 = np.arange(215, 500 + 0.5, 0.5) * wavelen0
    ZItotalDisplay = np.zeros([2 * nn + 2, Num])
    intensitymax = 0
    ItotalDisplaymax = np.zeros([2 * nn + 2, 2 * nn + 2])
    for nnz in range(Num):
        print(nnz)
        Ex1, Ey1, Ez1 = Diffraction2DTransPolar(Ex0, Ey0, Zd1[nnz], lam[wavelennum - 1], n_refra, Dx, samplenum, nn)
        ItotalDisplay = np.abs(Ex1) ** 2 + np.abs(Ey1) ** 2 + np.abs(Ez1) ** 2
        ZItotalDisplay[:, nnz] = ItotalDisplay[:, nn]
        if ItotalDisplay[nn, nn] > intensitymax:
            intensitymax = ItotalDisplay[nn, nn]
            ItotalDisplaymax = ItotalDisplay
    ItotalDisplay_incident = np.abs(Ex0) ** 2 + np.abs(Ey0) ** 2
    df3 = pd.DataFrame(ZItotalDisplay.get())
    filename = 'D:\\zhaofen\\huangbaoze\\A_gdgddfold\\otherwavelengths_verification\\fold_random_7\\ZItotalDisplay\\' + 'ZItotalDisplay' + str(wavelennum) + '.xlsx'
    df3.to_excel(filename, index=False, header=False)
    ratio1, distance1, strength1, SL, efficiency = computing_plot(Zd1, ZItotalDisplay, wavelennum, Num, nn, Dx, wavelen0, lam, ItotalDisplay_incident, ItotalDisplaymax)
    return ratio1, distance1, strength1, SL, efficiency


def Diffraction2DTransPolar(Ex0, Ey0, Z, wavelen, n_refr, Dx, samplenum, nn):

    num = np.arange(samplenum)
    freq = 1 / (samplenum * Dx) * (num - samplenum / 2 + 0.5)
    freq_x = np.outer(freq, np.ones(samplenum))
    freq_y = freq_x.T
    # freq_y = np.ones([samplenum, 1]) * freq
    # freq_x = freq_y.T
    fz = np.sqrt(((n_refr / wavelen) ** 2 - freq_x ** 2 - freq_y ** 2).astype(complex))
    SpectrumX = FourrierTrans2D(Ex0, Dx, samplenum, 1)
    SpectrumX = SpectrumX * np.exp(1j * 2 * np.pi * fz * Z)  #如果可以，可以试下Z为向量，会不会一下子就算出不同轴向距离的场分布情况？？？
    Ex = FourrierTrans2D(SpectrumX, Dx, samplenum, -1)
    Ex1 = Ex[(samplenum / 2 - 1 - nn): (samplenum / 2 + nn + 1), (samplenum / 2 - 1 - nn): (samplenum / 2 + nn + 1)]

    SpectrumY = FourrierTrans2D(Ey0, Dx, samplenum, 1)
    SpectrumY = SpectrumY * np.exp(1j * 2 * np.pi * fz * Z)
    Ey = FourrierTrans2D(SpectrumY, Dx, samplenum, -1)
    Ey1 = Ey[(samplenum / 2 - 1 - nn): (samplenum / 2 + nn + 1), (samplenum / 2 - 1 - nn): (samplenum / 2 + nn + 1)]

    SpectrumZ = -(freq_x * SpectrumX + freq_y * SpectrumY) / fz * np.exp(1j * 2 * np.pi * fz * Z)
    Ez = FourrierTrans2D(SpectrumZ, Dx, samplenum, -1)
    Ez1 = Ez[(samplenum / 2 - 1 - nn): (samplenum / 2 + nn + 1), (samplenum / 2 - 1 - nn): (samplenum / 2 + nn + 1)]

    return Ex1, Ey1, Ez1


def FourrierTrans2D(g , Dx, samplenum, flag):

    global G
    num = np.arange(samplenum)
    a = np.exp(1j * 2 * np.pi / samplenum * (samplenum / 2 - 0.5) * num)
    A = np.outer(a, a)
    C = np.exp(-1j * 2 * np.pi / samplenum * (samplenum / 2 - 0.5) ** 2 * 2) * A
    if flag == 1:
        G = Dx ** 2 * C * np.fft.fft2(A * g)
    if flag == -1:
        G = (1 / (samplenum * Dx)) ** 2 * samplenum ** 2 * np.conj(C) * np.fft.ifft2(np.conj(A) * g)
    return G

def computing_plot(Z, AllDisplay, wavelennum, Num, nn, Dx, wavelen0, lam, ItotalDisplay_incident, ItotalDisplaymax):

    XX = np.arange(-(nn * Dx + 0.5 * Dx), (nn * Dx + 0.5 * Dx), Dx)
    Ipeak = np.zeros(Num)
    for i in range(Num):
        Ipeak[i] = AllDisplay[nn, i]
    strength1 = np.max(Ipeak)
    mark = (Ipeak == strength1)
    distance1 = Z[mark] / wavelen0

    j = nn + 1
    while AllDisplay[j, mark] > strength1 / 2 and j < 2 * nn + 1:
        j = j + 1
    I1 = AllDisplay[j - 1, mark]
    x1 = XX[j - 1]
    I2 = AllDisplay[j, mark]
    x2 = XX[j]
    b = (I2 - I1) / (x2 - x1)
    c = I2 - b * x2
    FWHM1 = (0.5 * strength1 - c) / b

    k = nn - 1
    while AllDisplay[k, mark] > strength1 / 2 and k > 0:
        k = k - 1
    I1 = AllDisplay[k + 1, mark]
    x1 = XX[k + 1]
    I2 = AllDisplay[k, mark]
    x2 = XX[k]
    b = (I2 - I1) / (x2 - x1)
    c = I2 - b * x2
    FWHM2 = (0.5 * strength1 - c) / b

    ratio1 = (FWHM1 - FWHM2) / lam[wavelennum - 1]


    t = np.where(Ipeak == strength1)[0]
    while Ipeak[t] > 10 and t > 1:
        t -= 1
    while Ipeak[t + 1] >= Ipeak[t] >= Ipeak[t - 1] and t > 1:
        t -= 1
    SideLobe = np.max(Ipeak[0: t + 1])
    SL = SideLobe / strength1
    data = np.array([[SideLobe, strength1, SL]])
    file4 = 'D:\\zhaofen\\huangbaoze\\A_gdgddfold\\otherwavelengths_verification\\fold_random_7\\' + 'ZItotalDisplay_SideLobe_Imax_SL_ItotalDisplay_all.xlsx'
    df4 = pd.DataFrame(data.get())
    if wavelennum == 1:
        df4.to_excel(file4, sheet_name='Sheet1', startrow=int(wavelennum) - 1, header=False, index=False)
    else:
        book = load_workbook(file4)
        with pd.ExcelWriter(file4) as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df4.to_excel(writer, sheet_name='Sheet1', startrow=int(wavelennum) - 1, header=False, index=False)

    ItotalDisplay_incident_sum = sum(sum(ItotalDisplay_incident))
    xx = np.where(ItotalDisplaymax[nn, :] >= np.max(ItotalDisplaymax)/2)[0][-1] - nn
    x = np.where((ItotalDisplaymax / np.max(ItotalDisplaymax)) >= 1 / np.exp(1))[0]
    y = np.where((ItotalDisplaymax / np.max(ItotalDisplaymax)) >= 1 / np.exp(1))[1]
    ItotalDisplay_sum = 0
    for i in range(np.size(x)):
        if np.sqrt((x[i] - nn) ** 2 + (y[i] - nn) ** 2) > xx:
            tmp = 0
        else:
            tmp = ItotalDisplaymax[x[i], y[i]]
        ItotalDisplay_sum = ItotalDisplay_sum + tmp
    efficiency = ItotalDisplay_sum / ItotalDisplay_incident_sum
    data = np.array([[ItotalDisplay_sum, ItotalDisplay_incident_sum, efficiency]])
    df5 = pd.DataFrame(data.get())
    book = load_workbook(file4)
    with pd.ExcelWriter(file4) as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df5.to_excel(writer, sheet_name='Sheet1', startrow=int(wavelennum) - 1, startcol=4, header=False, index=False)

    return ratio1, distance1, strength1, SL, efficiency





































