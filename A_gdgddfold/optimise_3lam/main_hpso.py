try:
    import cupy as np
    print(f"cupy{np.__version__}")
except ModuleNotFoundError:
    print(ModuleNotFoundError)
    import numpy as np
import pandas as pd
from openpyxl import load_workbook
from fun_hpso import Fun_diffraction

class gd_gdd_fold:
    def __init__(self):
        self.lam = np.arange(8, 12 + 0.1, 0.1)
        self.lamc = self.lam[20]
        self.focallength = 295 * self.lamc #315 * self.lamc
        C = 2.99792458 * 10 ** 14 / 1e12  # um/ps 光速
        self.w0 = 2 * np.pi * C / self.lamc
        self.w = 2 * np.pi * C / self.lam
        self.f = self.focallength / self.w0 ** 2 * self.w ** 2
        self.T = 7.25
        self.N_part = 50
        self.R = 187 * self.lamc #设定半径
        self.N = np.floor((self.R - self.T / 2) / self.T)
        self.N = int(self.N)
        print(self.N)

        # self.fold_num = np.floor((self.N + 1) / (self.N_part + 1))
        # self.fold_num = int(self.fold_num)
        self.r = np.arange(0, (self.N + 1) * self.T, self.T)
        self.phi = 2 * np.pi / self.lamc * (self.focallength ** 2 - np.sqrt(self.r ** 2 + self.focallength ** 2))
        Num_L_W_A_P_beta = np.array(pd.read_excel('E:\\huangbaoze\\python\\A_gdgddfold\\Num_L_W_A_P_betarad.xlsx', header=None, sheet_name=0))
        s = np.shape(Num_L_W_A_P_beta)[0]
        nn = np.floor(self.phi / (2 * np.pi))
        self.phi = self.phi - nn * 2 * np.pi
        aa = 0
        self.GeneNum = np.zeros(self.N + 1)
        for i in range(self.N + 1):
            if self.phi[i] >= Num_L_W_A_P_beta[15, 4]:
                if np.abs(2 * np.pi - self.phi[i]) <= np.abs(self.phi[i] - Num_L_W_A_P_beta[15, 4]):
                    aa = 0
                else:
                    aa = 15
            else:
                for k in range(s - 1):
                    if Num_L_W_A_P_beta[k + 1, 4] > self.phi[i] >= Num_L_W_A_P_beta[k, 4]:
                        if np.abs(self.phi[i] - Num_L_W_A_P_beta[k, 4]) <= np.abs(
                                self.phi[i] - Num_L_W_A_P_beta[k + 1, 4]):
                            aa = k
                        else:
                            aa = k + 1
            self.phi[i] = Num_L_W_A_P_beta[aa, 4]
            self.GeneNum[i] = Num_L_W_A_P_beta[aa, 0]

        self.xx = np.ones((2 * self.N + 1, 1)) * np.arange(-self.N * self.T, (self.N + 1) * self.T, self.T)
        self.yy = (np.ones((2 * self.N + 1, 1)) * np.arange(self.N * self.T, -(self.N + 1) * self.T, -self.T)).T
        self.rr = np.sqrt(self.xx ** 2 + self.yy ** 2)
        self.bandnum = np.floor(self.rr / self.T) + 1
        self.R = (self.N + 0.5) * self.T
        self.bandnum[self.rr > self.R] = 0
        self.bandnum = self.bandnum.astype(int)
        self.samplenum = 2048

        self.Dx = 2 * self.R / self.samplenum
        self.lamnum = 41

        ##GD/GDD分布
        self.coefficient_matrix = np.array(pd.read_excel('E:\\huangbaoze\\python\\A_gdgddfold\\new_coefficient_matrix.xlsx', header=None, sheet_name=0))
        self.populationall = 5
        self.iterationall = 500
        self.wmax = 0.9
        self.wmin = 0.4
        self.c1 = 2
        self.c2 = 3


    def run(self):

        fold_GD = np.zeros(self.N + 1)
        fold_GDD = np.zeros(self.N + 1)
        fold_GD[:] = self.coefficient_matrix[3, 0: self.N + 1]
        fold_GDD[:] = self.coefficient_matrix[2, 0: self.N + 1]

        fold_num = 0
        flag = 1
        GD_fold_loc = self.N_part
        # GDD_fold_loc = self.N_part
        lac_num1 = np.zeros(0)

        global lac_num, fold_phase1, fold_phase
        fold_phi = np.zeros([self.populationall, self.N + 1])
        for k in range(self.populationall):
            fold_phi[k] = self.phi

        while flag:
            fold_num += 1
            last_GD_fold_loc = GD_fold_loc
            # last_GDD_fold_loc = GDD_fold_loc

            GD_fold_loc = last_GD_fold_loc + np.where((self.coefficient_matrix[3, last_GD_fold_loc + 1:] - self.coefficient_matrix[3, last_GD_fold_loc + 1])
                                                <= self.coefficient_matrix[3, self.N_part])[0][-1] + 1
            # GDD_fold_loc = last_GDD_fold_loc + np.where(np.abs(self.coefficient_matrix[2, last_GDD_fold_loc + 1:] - self.coefficient_matrix[2, last_GDD_fold_loc + 1])
            #                                     <= self.coefficient_matrix[2, self.N_part])[0][-1] + 1
            # if GD_fold_loc <= GDD_fold_loc:
            #     GDD_fold_loc = GD_fold_loc
            # else:
            #     GD_fold_loc = GDD_fold_loc

            lac_num = np.zeros([1, fold_num])
            lac_num[0] = np.append(lac_num1, last_GD_fold_loc)
            lac_num1 = lac_num
            fold_GD[last_GD_fold_loc + 1: GD_fold_loc + 1] = self.coefficient_matrix[3, last_GD_fold_loc + 1: GD_fold_loc + 1] - \
                                                          self.coefficient_matrix[3, last_GD_fold_loc + 1]
            fold_GDD[last_GD_fold_loc + 1: GD_fold_loc + 1] = self.coefficient_matrix[2, last_GD_fold_loc + 1: GD_fold_loc + 1] - \
                                                         self.coefficient_matrix[2, last_GD_fold_loc + 1]

            # fold_phase_tmp = np.random.rand(self.populationall, 1) * 2 * np.pi
            # # fold_phase_tmp = np.zeros([self.populationall, 1]) #相位不连续为0
            #
            # if fold_num == 1:
            #     fold_phase1 = np.zeros([self.populationall, 0])
            #
            # fold_phase = np.zeros([self.populationall, fold_num])
            # for s in range(self.populationall):
            #     fold_phi[s, last_GD_fold_loc + 1: GD_fold_loc + 1] = self.phi[last_GD_fold_loc + 1: GD_fold_loc + 1] + fold_phase_tmp[s, 0]
            #     fold_phase[s] = np.append(fold_phase1[s], fold_phase_tmp[s, 0])
            # fold_phase1 = fold_phase

            if GD_fold_loc == self.N:# or GDD_fold_loc == self.N:
                flag = 0
        fileall = 'E:\\huangbaoze\\python\\A_gdgddfold\\optimise_3lam\\new_fold\\'
        df_fold_GD = pd.DataFrame(fold_GD.get())
        df_fold_GDD = pd.DataFrame(fold_GDD.get())
        # # df_fold_phase = pd.DataFrame(fold_phase1.get())
        df_fold_GD.to_excel(fileall + 'df_fold_GD.xlsx', sheet_name='Sheet1', startrow=0, header=False, index=False)
        df_fold_GDD.to_excel(fileall + 'df_fold_GDD.xlsx', sheet_name='Sheet1', startrow=0, header=False, index=False)
        # # df_fold_phase.to_excel(fileall + 'df_fold_phase.xlsx', sheet_name='Sheet1', startrow=0, header=False, index=False)
        # print(lac_num1.get())
        df_lac_num = pd.DataFrame(lac_num.get())
        df_lac_num.to_excel(fileall + 'df_lac_num.xlsx', sheet_name='Sheet1', startrow=0, header=False, index=False)
        # df1 = pd.DataFrame(fold_phi.get())
        # df1.to_excel(fileall + 'fold_phi.xlsx', sheet_name='Sheet1', startrow=0, header=False, index=False)
        lamnum = np.array([1, 11, 21, 31, 41])
        num = 5
        # lamnum = np.arange(41) + 1
        # num = 41

        matrixband = np.zeros([num, 2 * self.N + 1, 2 * self.N + 1])
        matrixPhaseBandGap = np.zeros([num, 2 * self.N + 1, 2 * self.N + 1])
        for m in range(num):
            for n in range(self.N + 1):
                matrixband[m, self.bandnum == n + 1] = 1
                matrixPhaseBandGap[m, self.bandnum == n + 1] = fold_GD[n] * (self.w[lamnum[m] - 1]
                                                                             - self.w0) + fold_GDD[n] * (
                                                                       self.w[lamnum[m] - 1] - self.w0) ** 2
            # df2 = pd.DataFrame(matrixband[m].get())
            # df3 = pd.DataFrame(matrixPhaseBandGap[m].get())
            # if m == 0:
            #     df2.to_excel(fileall + 'matrixband.xlsx', sheet_name=str(m), header=False, index=False)
            #     df3.to_excel(fileall + 'matrixPhaseBandGap.xlsx', sheet_name=str(m), header=False, index=False)
            # else:
            #     with pd.ExcelWriter(fileall + 'matrixband.xlsx', mode='a') as writer:
            #         df2.to_excel(writer, sheet_name=str(m), header=False, index=False)
            # with pd.ExcelWriter(fileall + 'matrixPhaseBandGap.xlsx', mode='a') as writer:
            #     df3.to_excel(writer, sheet_name=str(m), header=False, index=False)

        min_fold_phase = np.zeros([1, fold_num])
        min_fold_fitness = np.zeros(1)
        min_Distance = np.zeros(1)
        min_Strength = np.zeros(1)
        min_FWHM = np.zeros(1)
        min_Dlimit = np.zeros(1)
        min_sl = np.zeros(1)
        min_Efficiency = np.zeros(1)
        min_lam = np.zeros(1)

        record_fold_phase = np.zeros([self.populationall, fold_num])
        record_fold_fitness = np.zeros(self.populationall)
        record_Distance = np.zeros(self.populationall)
        record_Strength = np.zeros(self.populationall)
        record_FWHM = np.zeros(self.populationall)
        record_Dlimit = np.zeros(self.populationall)
        record_sl = np.zeros(self.populationall)
        record_Efficiency = np.zeros(self.populationall)
        record_lam = np.zeros(self.populationall)

        fold_fitness = np.zeros(self.populationall)
        Distance = np.zeros(self.populationall)
        Strength = np.zeros(self.populationall)
        FWHM = np.zeros(self.populationall)
        Dlimit = np.zeros(self.populationall)
        sl = np.zeros(self.populationall)
        Efficiency = np.zeros(self.populationall)
        lam = np.zeros(self.populationall)

        v = np.random.rand(self.populationall, fold_num) / 50 -0.01 #-0.01~0.01
        for iteration in range(self.iterationall):
            print(iteration)
            if iteration == 0:
                fold_phase = np.random.rand(self.populationall, fold_num) * 2 * np.pi #初始化中心波长对应相位不连续
                # fold_phase = np.zeros([self.populationall, fold_num])  # 中心波长对应相位不连续为0

            for l in range(self.populationall):
                if iteration > 0:
                    w = self.wmax - (self.wmax - self.wmin) / (self.iterationall ** 2) * iteration ** 2
                    v[l] = w * v[l] + self.c1 * np.random.rand(1) * (record_fold_phase[l] - fold_phase[l]) + self.c2 * \
                           np.random.rand(1) * (min_fold_phase[0] - fold_phase[l])

                    fold_phase[l] = fold_phase[l] + v[l]

                    tmpv1 = v[l] > 0.1
                    v[l, tmpv1] = 0.1
                    tmpv2 = v[l] < -0.1
                    v[l, tmpv2] = -0.1

                    tmpphase1 = fold_phase[l] > 2 * np.pi
                    # fold_phase[l, tmpphase1] = 2 * np.pi
                    mm1 = np.floor(fold_phase[l, tmpphase1] / (2 * np.pi))
                    fold_phase[l, tmpphase1] = fold_phase[l, tmpphase1] - mm1 * 2 * np.pi #破壁
                    tmpphase2 = fold_phase[l] < 0
                    # fold_phase[l, tmpphase2] = 0
                    mm2 = np.floor(fold_phase[l, tmpphase2] / (2 * np.pi))
                    fold_phase[l, tmpphase2] = fold_phase[l, tmpphase2] - mm2 * 2 * np.pi #破壁


                for t in range(fold_num):
                    fold_phi[l, int(lac_num[0, t]) + 1:] = self.phi[int(lac_num[0, t]) + 1:] + fold_phase[l, t]

                for j in range(num):
                    ##角谱衍射
                    ratio1, distance1, strength1, SL, efficiency = Fun_diffraction(
                        fold_phi[l], matrixband[j], matrixPhaseBandGap[j], lamnum[j], self.samplenum, self.Dx, self.T,
                        self.N, self.bandnum, self.lamc, self.lam, self.f / self.lamc)
                    print(distance1)
                    # fitness = np.abs(distance1 - self.f[int(lamnum[j] - 1)] / self.lamc) / (10 ** j)
                    if j == 0:
                        fitness = np.abs(distance1 - self.f[int(lamnum[j] - 1)] / self.lamc)
                        fold_fitness[l] = fitness
                        Distance[l] = distance1
                        Strength[l] = strength1
                        FWHM[l] = ratio1
                        DLWavelen = 0.5 / np.sin(np.arctan(self.R / (distance1 * self.lamc)))
                        Dlimit[l] = DLWavelen
                        sl[l] = SL
                        Efficiency[l] = efficiency
                        lam[l] = j + 1
                    else:
                        if np.abs(distance1 - self.f[int(lamnum[j] - 1)] / self.lamc) > fitness:
                            fitness = np.abs(distance1 - self.f[int(lamnum[j] - 1)] / self.lamc)
                            fold_fitness[l] = fitness
                            Distance[l] = distance1
                            Strength[l] = strength1
                            FWHM[l] = ratio1
                            DLWavelen = 0.5 / np.sin(np.arctan(self.R / (distance1 * self.lamc)))
                            Dlimit[l] = DLWavelen
                            sl[l] = SL
                            Efficiency[l] = efficiency
                            lam[l] = j + 1

                    # if j == 0 and fitness > 10:
                    #      break
                    # if j == 0 and fitness < 10:
                    #     continue
                    # if j == 1 and fitness > 1:
                    #     break
                    # if j == 1 and fitness < 1:
                    #     continue

                if iteration > 0:
                    if fold_fitness[l] < record_fold_fitness[l]:
                        record_fold_fitness[l] = fold_fitness[l]
                        record_Distance[l] = Distance[l]
                        record_Strength[l] = Strength[l]
                        record_FWHM[l] = FWHM[l]
                        record_Dlimit[l] = Dlimit[l]
                        record_sl[l] = sl[l]
                        record_Efficiency[l] = Efficiency[l]
                        record_fold_phase[l] = fold_phase[l]
                        record_lam[l] = lam[l]

                    if record_fold_fitness[l] < min_fold_fitness[0]:
                        min_fold_fitness[0] = record_fold_fitness[l]
                        min_Distance[0] = record_Distance[l]
                        min_Strength[0] = record_Strength[l]
                        min_FWHM[0] = record_FWHM[l]
                        min_Dlimit[0] = record_Dlimit[l]
                        min_sl[0] = record_sl[l]
                        min_Efficiency[0] = record_Efficiency[l]
                        min_fold_phase[0] = record_fold_phase[l]
                        min_lam[0] = record_lam[l]

            if iteration == 0:
                for p in range(self.populationall):
                    record_fold_phase[p] = fold_phase[p]
                    record_fold_fitness[p] = fold_fitness[p]
                    record_Distance[p] = Distance[p]
                    record_Strength[p] = Strength[p]
                    record_FWHM[p] = FWHM[p]
                    record_Dlimit[p] = Dlimit[p]
                    record_sl[p] = sl[p]
                    record_Efficiency[p] = Efficiency[p]
                    record_lam[p] = lam[p]

                location = np.where(record_fold_fitness == np.min(record_fold_fitness))[0]
                min_fold_phase[0] = record_fold_phase[location[0]]
                min_fold_fitness[0] = record_fold_fitness[location[0]]
                min_Distance[0] = record_Distance[location[0]]
                min_Strength[0] = record_Strength[location[0]]
                min_FWHM[0] = record_FWHM[location[0]]
                min_Dlimit[0] = record_Dlimit[location[0]]
                min_sl[0] = record_sl[location[0]]
                min_Efficiency[0] = record_Efficiency[location[0]]
                min_lam[0] = record_lam[location[0]]

            data = np.array([[min_fold_fitness[0], min_Distance[0], min_Strength[0], min_FWHM[0], min_Dlimit[0], min_sl[0], min_Efficiency[0], min_lam[0]]])
            print(data.get())
            dfdata1 = pd.DataFrame(data.get())
            dfdata2 = pd.DataFrame(min_fold_phase[0].reshape([1,-1]).get())
            global sign
            filedata1 = 'E:\\huangbaoze\\python\\A_gdgddfold\\optimise_3lam\\data\\' + str(sign) + '_min_fold_fitness.xlsx'
            filedata2 = 'E:\\huangbaoze\\python\\A_gdgddfold\\optimise_3lam\\data\\' + str(sign) + '_min_fold_phase.xlsx'
            if iteration == 0:
                dfdata1.to_excel(filedata1, sheet_name='Sheet1', startrow=iteration, header=False, index=False)
                dfdata2.to_excel(filedata2, sheet_name='Sheet1', startrow=iteration, header=False, index=False)
            else:
                book = load_workbook(filedata1)
                with pd.ExcelWriter(filedata1) as writer:
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    dfdata1.to_excel(writer, sheet_name='Sheet1', startrow=iteration, header=False, index=False)
                book = load_workbook(filedata2)
                with pd.ExcelWriter(filedata2) as writer:
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    dfdata2.to_excel(writer, sheet_name='Sheet1', startrow=iteration, header=False, index=False)
                # with pd.ExcelWriter(filedata2, mode='a') as writer:
                #     dfdata2.to_excel(writer, sheet_name=str(iteration), header=False, index=False)

        return min_fold_fitness[0]

sign = 20
if __name__ == '__main__':
    main = gd_gdd_fold()
    kk = main.run()
    while kk > 10:
        sign += 1
        main = gd_gdd_fold()
        kk = main.run()










